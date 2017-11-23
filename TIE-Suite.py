import vlc
import os
import sys
from Tkinter import *
import ttk
from tkFileDialog import *
# import pathlib
from threading import Thread, Event
import time
import platform
from openpyxl import *
# from datetime import datetime
import pathlib
from __builtin__ import file
import tkFileDialog
from openpyxl.styles.borders import Side

gridi = 0
gridj = 0
excelFile = ""
themeFile = ""
themeFile2 = ""
vidFile = ""
buttonDict = {}
themeBuild = {}


class ttkTimer(Thread):
    """a class serving same function as wxTimer... but there may be better ways to do this
    """

    def __init__(self, callback, tick):
        Thread.__init__(self)
        self.callback = callback
        self.stopFlag = Event()
        self.tick = tick
        self.iters = 0

    def run(self):
        while not self.stopFlag.wait(self.tick):
            self.iters += 1
            self.callback()

    def stop(self):
        self.stopFlag.set()

    def get(self):
        return self.iters


    # Borrowed/Adapted from one of the open-source examples bundled with python-vlc
class Player(Frame):
    """The main window has to deal with events.
    """

    def __init__(self, parent, title=None):
        Frame.__init__(self, parent)

        self.parent = parent
        loadT1 = StringVar()
        loadT2 = StringVar()
        loadW = StringVar()
        loadV = StringVar()
        
        
        if title == None:
            title = "Tie Sweet 1.0"
        self.parent.title(title)

        # Menu Bar
        #   File Menu
        menubar = Menu(self.parent)
        self.parent.config(menu=menubar)

        # Menu for running suite
        runMenu = Menu(menubar, tearoff=0)
        runMenu.add_command(label="Play Video", underline=0, command=self.OnOpen)
        runMenu.add_command(label="Launch TIE Coder - Client", command=lambda: playerInterface(self, "Client"))
        runMenu.add_command(label="Launch TIE Coder - Counselor", command=lambda: playerInterface(self, "Counselor"))
        runMenu.add_separator()
        runMenu.add_command(label="Save and Quit", command=lambda: saveQuit(self))
        menubar.add_cascade(label="Run", menu=runMenu)

        # Create projMenu Menu List
        projMenu = Menu(menubar, tearoff=0)
        projMenu.add_command(label="Resume Project", command=lambda: resume(parent, loadT1, loadT2, loadW, loadV))
        projMenu.add_command(label="Load Project", command=lambda: loadProj(parent, loadT1, loadT2, loadW, loadV))
        projMenu.add_command(label="Save Project", command=lambda: saveProj(parent))
        menubar.add_cascade(label="Project", menu=projMenu)

        # Load in existing files
        loadMenu = Menu(menubar, tearoff=0)
        loadMenu.add_command(label="Load Theme - Client", command=lambda: loadTheme(parent, "Client", loadT1))
        loadMenu.add_command(label="Load Theme - Counselor", command=lambda: loadTheme(parent, "Counselor", loadT2))
        loadMenu.add_command(label="Load Workbook", command=lambda: loadExcel(parent, loadW))
        loadMenu.add_command(label="Load Video File", command=lambda: vidLoad(parent, loadV))
        menubar.add_cascade(label="Load Files", menu=loadMenu)
        
        # Create new files
        newMenu = Menu(menubar, tearoff=0)
        newMenu.add_command(label="Create New Theme", command=lambda: importTheme(parent))
        newMenu.add_command(label="Create New Workbook", command=lambda: createExcel(parent))
        menubar.add_cascade(label="Create Files", menu=newMenu)
        
        #Show the User which Files are loaded
        self.load1 = Label(self.parent, textvariable=loadT1).pack()
        self.load2 = Label(self.parent, textvariable=loadT2).pack()
        self.load3 = Label(self.parent, textvariable=loadW).pack()
        self.load4 = Label(self.parent, textvariable=loadV).pack()
        
        
        # The second panel holds controls
        self.player = None
        self.videopanel = ttk.Frame(self.parent)
        self.canvas = Canvas(self.videopanel).pack(fill=BOTH, expand=1)
        self.videopanel.pack(fill=BOTH, expand=1)

        ctrlpanel = ttk.Frame(self.parent)
        l = Label(ctrlpanel, text="                         ")
        pause = ttk.Button(ctrlpanel, text="Pause", command=self.OnPause)
        play = ttk.Button(ctrlpanel, text="Play", command=self.OnPlay)
        stop = ttk.Button(ctrlpanel, text="Stop", command=self.OnStop)
        volume = ttk.Button(ctrlpanel, text="Mute", command=self.OnToggleVolume)
        pause.pack(side=LEFT)
        play.pack(side=LEFT)
        stop.pack(side=LEFT)
        l.pack(side=LEFT)
        volume.pack(side=LEFT)
        self.volume_var = IntVar()
        self.volslider = Scale(ctrlpanel, variable=self.volume_var, command=self.volume_sel,
                from_=0, to=100, orient=HORIZONTAL, length=100)
        self.volslider.pack(side=RIGHT)
        ctrlpanel.pack(side=BOTTOM)
        
        l1 = Label(self.parent, text="")
        l1.pack(side=BOTTOM)

        ctrlpanel2 = ttk.Frame(self.parent)
        self.scale_var = DoubleVar()
        self.timeslider_last_val = ""
        self.timeslider = Scale(ctrlpanel2, variable=self.scale_var, command=self.scale_sel,
                from_=0, to=10000, orient=HORIZONTAL, length=500)
        self.timeslider.pack(side=BOTTOM, fill=X, expand=1)
        self.timeslider_last_update = time.time()
        ctrlpanel2.pack(side=BOTTOM, fill=X)

        # VLC player controls
        self.Instance = vlc.Instance()
        self.player = self.Instance.media_player_new()
        self.timer = ttkTimer(self.OnTimer, 1.0)
        self.timer.start()
        self.parent.update()

        # self.player.set_hwnd(self.GetHandle()) # for windows, OnOpen does does this

    def OnExit(self, evt):
        """Closes the window.
        """
        self.Close()

    def OnOpen(self):
        """Pop up a new dialow window to choose a file, then play the selected file.
        """
        # If a file is already running, then stop it.
        self.OnStop()

        # Setting the media to the vidFile from the project
        self.Media = self.Instance.media_new(vidFile)
        self.player.set_media(self.Media)

        # Set the window id where to render VLC's video output
        if platform.system() == 'Windows':
            self.player.set_hwnd(self.GetHandle())
            # print("win")
        else:
            self.player.set_xwindow(self.GetHandle())  # this line messes up windows
            # print("win")

        self.OnPlay()
        # self.player.audio_set_volume(100)
            # set the volume slider to the current volume
            # self.volslider.SetValue(self.player.audio_get_volume() / 2)
        self.volslider.set(self.player.audio_get_volume())

    def OnPlay(self):
        """Toggle the status to Play/Pause.
        If no file is loaded, open the dialog window.
        """
        # check if there is a file to play, otherwise open a
        # Tk.FileDialog to select a file
        if not self.player.get_media():
            self.OnOpen()
        else:
            # Try to launch the media, if this fails display an error message
            if self.player.play() == -1:
                self.errorDialog("Unable to play.")

    def GetHandle(self):
        return self.videopanel.winfo_id()

    # def OnPause(self, evt):
    def OnPause(self):
        """Pause the player.
        """
        self.player.pause()

    def OnStop(self):
        """Stop the player.
        """
        self.player.stop()
        # reset the time slider
        self.timeslider.set(0)

    def OnTimer(self):
        """Update the time slider according to the current movie time.
        """
        if self.player == None:
            return
        # since the self.player.get_length can change while playing,
        # re-set the timeslider to the correct range.
        length = self.player.get_length()
        dbl = length * 0.001
        self.timeslider.config(to=dbl)

        # update the time on the slider
        tyme = self.player.get_time()
        if tyme == -1:
            tyme = 0
        dbl = tyme * 0.001
        self.timeslider_last_val = ("%.0f" % dbl) + ".0"
        # don't want to programatically change slider while user is messing with it.
        # wait 2 seconds after user lets go of slider
        if time.time() > (self.timeslider_last_update + 2.0):
            self.timeslider.set(dbl)

    def scale_sel(self, evt):
        if self.player == None:
            return
        nval = self.scale_var.get()
        sval = str(nval)
        if self.timeslider_last_val != sval:
            self.timeslider_last_update = time.time()
            mval = "%.0f" % (nval * 1000)
            self.player.set_time(int(mval))  # expects milliseconds

    def volume_sel(self, evt):
        if self.player == None:
            return
        volume = self.volume_var.get()
        if volume > 100:
            volume = 100
        if self.player.audio_set_volume(volume) == -1:
            self.errorDialog("Failed to set volume")

    def OnToggleVolume(self):
        """Mute/Unmute according to the audio button.
        """
        is_mute = self.player.audio_get_mute()

        self.player.audio_set_mute(not is_mute)
        # update the volume slider;
        # since vlc volume range is in [0, 200],
        # and our volume slider has range [0, 100], just divide by 2.
        self.volume_var.set(self.player.audio_get_volume())

    def OnSetVolume(self):
        """Set the volume according to the volume sider.
        """
        volume = self.volume_var.get()
        # vlc.MediaPlayer.audio_set_volume returns 0 if success, -1 otherwise
        if volume > 100:
            volume = 100
        if self.player.audio_set_volume(volume) == -1:
            self.errorDialog("Failed to set volume")

    def errorDialog(self, errormessage):
        """Display a simple error dialog"""
        Tk.tkMessageBox.showerror(self, 'Error', errormessage)

    def get_time(self):
        return self.player.get_time()


    # Button class, contains data and write to excel functionality
class but:

    def __init__ (self, key, name, type, master, player, observed):
        self.key = key
        self.name = name
        self.type = type
        self.active = 0
        self.time = ""
        self.time2 = ""
        self.time3 = 0
        self.time4 = 0
        self.player = player
        self.observed = observed
        global gridi
        global gridj
        global excelFile
        #print(self.type)
        if self.type == "1":
            b = Button(
                master, text = key, command = lambda: self.createEntry(excelFile,self.time, self.time),
                height = 2, width =  4
                )
        else:
            b = Checkbutton(
                master, text = key, command = lambda: self.createEntry(excelFile,self.time, self.time),
                height = 2, width = 4, indicatoron = 0
                )

        b.grid(row = gridi, column = (gridj%10))
        #print(gridi)
        gridj += 1
        if (gridj%10) == 0:
            gridi += 1

    def createEntry(self, file, time, time2):
        if self.type == "1":
            wb = load_workbook(file)
            ws = wb.active
            index = ws.max_row + 1
            ws.cell(row = index, column = 1).value = self.observed
            ws.cell(row = index, column = 2).value = self.key
            ws.cell(row = index, column = 3).value = self.name
            ws.cell(row = index, column = 4).value = getTime(self.player, 1)
            ws.cell(row = index, column = 5).value = ""
            ws.cell(row = index, column = 6).value = getTime(self.player, 2)
            ws.cell(row = index, column = 7).value = ""
            wb.save(file)
        else:
            if self.active == 0:
                self.time = getTime(self.player,1)
                self.time3 = getTime(self.player,2)
                self.active = 1
            else:
                self.time2 = getTime(self.player,1)
                self.time4 = getTime(self.player,2)
                self.active = 0
                wb = load_workbook(file)
                ws = wb.active
                index = ws.max_row + 1
                ws.cell(row = index, column = 1).value = self.observed
                ws.cell(row = index, column = 2).value = self.key
                ws.cell(row = index, column = 3).value = self.name
                ws.cell(row = index, column = 4).value = self.time
                ws.cell(row = index, column = 5).value = self.time2
                ws.cell(row = index, column = 6).value = self.time3
                ws.cell(row = index, column = 7).value = self.time4
                ws.cell(row = index, column = 8).value = (self.time4 - self.time3)
                wb.save(file)


    # Creates the buttons used by the TIE Coder
class themeBut:

    def __init__ (self, key, name, type, master):
        self.key = key
        self.name = name
        self.type = type
        self.used = 0
        global gridi
        global gridj
        b = Checkbutton(
            master, text=key, command=lambda: self.change(),
            height=2, width=4, indicatoron=0
            )
        b.grid(row=gridi, column=(gridj % 10))
        gridj += 1
        if (gridj % 10) == 0:
            gridi += 1

    def change(self):
        if self.used == 0:
            self.used = 1
        else:
            self.used = 0

    def getUsed(self):
        return self.used


    # Used to get different time points in the playuer
def getTime(player, type):
    time = player.get_time()
    t = int(time / 1000)
    min = int(t / 60)
    if min < 10:
        mins = "0" + str(min)
    else:
        mins = str(min)
    sec = int(t % 60)
    if sec < 10:
        secs = "0" + str(sec)
    else:
        secs = str(sec)
    # print(min + " " + sec)
    timeStr = "" + mins + ":" + secs
    if type == 1:
        return timeStr
    else:
        return t


def Tk_get_root():
    if not hasattr(Tk_get_root, "root"):  # (1)
        Tk_get_root.root = Tk()  # initialization call is inside the function
    return Tk_get_root.root


    # Basic quit function
def _quit():
    # print("_quit: bye")
    root = Tk_get_root()
    root.quit()  # stops mainloop
    root.destroy()  # this is necessary on Windows to prevent
                    # Fatal Python Error: PyEval_RestoreThread: NULL tstate
    os._exit(1)


def saveQuit(master):
    
    global themeFile
    global themeFile2
    global excelFile
    global vidFile
    
    newRes = themeFile + ";" + themeFile2 + ";" + excelFile + ";" + vidFile
    file = open("./Resources/resumeFile.txt", 'w')
    file.write(newRes)
    _quit()


def resume(master, file1, file2, file3, file4):
    
    global themeFile
    global themeFile2
    global excelFile
    global vidFile
    
    projFile = "./Resources/resumeFile.txt"
    
    with open("./Resources/resumeFile.txt", "rU") as FILE:
        for line in FILE:
            line = line.strip()
            themeFile = line.split(";")[0]
            themeFile2 = line.split(";")[1]
            excelFile = line.split(";")[2]
            vidFile = line.split(";")[3]
    trim = os.path.basename(themeFile)
    file1.set("Client Theme: " + trim)
    trim = os.path.basename(themeFile2)
    file2.set("Counsellor Theme: " + trim)
    trim = os.path.basename(excelFile)
    file3.set("Workbook: " + trim)
    trim = os.path.basename(vidFile)
    file4.set("Video File: " + trim)        

    # Read a project file and load sub-files into memory
def loadProj(master, file1, file2, file3, file4):
    projFile = askopenfilename(initialdir="./Projects", title="Select a Project")

    global themeFile
    global themeFile2
    global excelFile
    global vidFile

    with open(projFile, "rU") as FILE:
        for line in FILE:
            line = line.strip()
            themeFile = line.split(";")[0]
            themeFile2 = line.split(";")[1]
            excelFile = line.split(";")[2]
            vidFile = line.split(";")[3]
    
    trim = os.path.basename(themeFile)
    file1.set("Client Theme: " + trim)
    trim = os.path.basename(themeFile2)
    file2.set("Counsellor Theme: " + trim)
    trim = os.path.basename(excelFile)
    file3.set("Workbook: " + trim)
    trim = os.path.basename(vidFile)
    file4.set("Video File: " + trim)

    # Query user for project name and pass name to projSave
def saveProj(master):

    pop = Tk()
    pop.title("Create a new Project File")
    frame = Frame(pop, height=225, width=300)
    frame.pack_propagate(0)
    frame.pack()

    l = Label(frame, text="To save a new project, make sure you have loaded:\n \n A Client Theme \n A Counsellor Theme \n A Workbook\n A Video File. \n \n Project File Name \n")
    l.pack()

    # Entry Window
    e = Entry(frame)
    e.pack()
    e.focus_set()
    
    l2 = Label(frame, text="")
    l2.pack()

    # Submission Button takes contents of Entry Window and calls newExcel on it
    b = ttk.Button(frame, text="Create Project", command=lambda: projSave("" + e.get(), pop))
    b.pack()


    # Save currently loaded files into project file based on input for later. 
def projSave(input, pop):
    global themeFile
    global themeFile2
    global excelFile
    global vidFile

    filename = "./Projects/" + input + ".txt"
    file = open(filename, "w")
    file.write(themeFile + ";" + themeFile2 + ";" + excelFile + ";" + vidFile)

    pop.destroy()


    # Tkinter interface for TIE Coder -- THIS NEEDS TO BE RENAMED
def playerInterface(player, observed):
    pop = Tk()
    pop.title("SweetTie Coder 1.0 - " + observed)
    global themeFile
    global themeFile2
    theme = {}
    if observed == "Client":
        theme = themeRead(themeFile)
    else:
        theme = themeRead(themeFile2)

    frame = Frame(pop)
    frame.pack_propagate(0)
    frame.pack()
    # Setting up button Widgets for container for GUI
    for (key, items) in sorted(theme.iteritems(), key=lambda (x, y):float(x)):
        but(items[0], items[1], items[2], frame, player, observed)


    # Reads a theme file and parses it into a dictionary
    # to be used to construct the buttons
def themeLine(line, theme):
    line = line.strip()
    theme[line.split(",")[0]] = [line.split(",")[1]]
    theme[line.split(",")[0]].append(line.split(",")[2])
    theme[line.split(",")[0]].append(line.split(",")[3])
    return theme


    # Parses the Theme's file into lines to be passed to themeLine for dictionary writing
def themeRead(file):
    theme = {}
    with open(file, "rU") as FILE:
        for line in FILE:
            themeLine(line, theme)
    return theme


    # Loads themes into memory for Client/Counsellor
def loadTheme(master, observed, file):
    # Querying user for theme.txt file
    global themeFile
    global themeFile2
    if observed == "Client":
        themeFile = tkFileDialog.askopenfilename(initialdir="./Themes", title="Select a Theme for the Client")
        trim = os.path.basename(themeFile)
        file.set("Client Theme: " + trim)
    else:
        themeFile2 = tkFileDialog.askopenfilename(initialdir="./Themes", title="Select a Theme for the Counsellor")
        trim = os.path.basename(themeFile)
        file.set("Counsellor Theme: " + trim)


    # Uses Master List to generate new custom themes
def importTheme(master):
    global themeBuild
    themeBuildFile = "./Resources/Theme Master List.txt"
    themeBuild = themeRead(themeBuildFile)
    buildTheme()


    # Builds the themeBuild dictionary to be used by newTheme
def buildTheme():
    global themeBuild
    global gridi
    global gridj
    global buttonDict
    gridi = 0
    gridj = 0
    
    # Setting up button container Widget for GUI
    pop1 = Tk()
    pop1.title("Creating New Theme")

    l = Label(pop1, text="\n Pick Existing Behaviors You'd like to Use \n")
    l.pack()
    
    frame = Frame(pop1, height=450, width=450)
    frame.pack_propagate(0)
    frame.pack()
    
    # Setting up button Widgets for container for GUI
    for (key, items) in sorted(themeBuild.iteritems(), key=lambda (x, y):float(x)):
        buttonDict[key] = themeBut(items[0], items[1], items[2], frame)
    
    l1 = Label(pop1, text="\n Please Name Your New Theme \n")
    l1.pack()
    
    e = Entry(pop1)
    e.pack()
    e.focus_set()
    
    l2 = Label(pop1, text=" ")
    l2.pack()
    
    b1 = ttk.Button(pop1, text="Finish", command=lambda: newTheme(e.get(), pop1))
    b1.pack()


    # Writes the themeBuild theme into a new Theme file
def newTheme(input, pop3):
    global themeBuild
    global buttonDict
    for (key, but) in buttonDict.items():
        if but.getUsed() == 0:
            del themeBuild[key]
    
    filename = "./Themes/" + input + ".txt"
    file = open(filename, "w")
    for (key, items) in themeBuild.items():
        line = "" + key + "," + items[0] + "," + items[1] + "," + items[2] + "\n"
        file.write(line)
        
    pop3.destroy()


    # Loads a given workbook into memory
def loadExcel(master, file):
    global excelFile
    excelFile = tkFileDialog.askopenfilename(initialdir="./Workbooks", title="Select a Workbook")
    trim = os.path.basename(excelFile)
    file.set("Workbook: " + trim)

    # Creates newExcel string for the init function 
def newExcel(input, pop):
    # Generating Blank Workbook
    wb = Workbook()
    global excelFile
    excelFile = "./Workbooks/" + input + ".xlsx"
    # Creating new Excel File
    wb.save(excelFile)
    # Initializing new Excel File
    init(excelFile)
    pop.destroy()
    # Confirmation of File Creation


    # Takes user input and passes to newExcel to make a new workbook based on input
def createExcel(master):
    global excelFile

    # Generating new Pop-Up Window
    pop = Tk()
    pop.title("Create a Workbook File")
    frame = Frame(pop, height=100, width=300)
    frame.pack_propagate(0)
    frame.pack()

    # Labelling intended content of the Entry Window
    l = Label(frame, text="New File Name")
    l.pack()

    # Entry Window
    e = Entry(frame)
    e.pack()
    e.focus_set()
    
    l1 = Label(frame, text="")
    l1.pack()

    # Submission Button takes contents of Entry Window and calls newExcel on it
    b = Button(frame, text="Create Workbook", width=15, command=lambda: newExcel(e.get(), pop))
    b.pack()


    # Initializes a new Workbook and Writes in coloumn headers. SHould prob rename this.
def init(file):
    wb = load_workbook(file)
    ws = wb.active
    ws.cell(row = 2, column = 1).value = "Observed"
    ws.column_dimensions["A"].width = 20
    ws.cell(row = 2, column = 2).value = "ID"
    ws.column_dimensions["B"].width = 8
    ws.cell(row = 2, column = 3).value = "Behavior Name"
    ws.column_dimensions["C"].width = 40
    ws.cell(row = 2, column = 4).value = "Time of Onset"
    ws.column_dimensions["D"].width = 15
    ws.cell(row = 2, column = 5).value = "Time of Offset"
    ws.column_dimensions["E"].width = 15
    ws.cell(row = 2, column = 6).value = "Onset in Secs"
    ws.column_dimensions["F"].width = 15
    ws.cell(row = 2, column = 7).value = "Offset in Secs"
    ws.column_dimensions["G"].width = 15
    ws.cell(row = 2, column = 8).value = "Duration in Secs"
    ws.column_dimensions["H"].width = 15
    wb.save(file)


    # Creates the vidFile string for the player
def vidLoad(master, file):
    global vidFile
    # p = pathlib.Path(os.path.expanduser("~"))
    vidFile = askopenfilename(initialdir="./Video Files", title="Choose a Video File", filetypes=(("all files", "*.*"), ("mp4 files", "*.mp4")))
    trim = os.path.basename(vidFile)
    file.set("Video File: " + trim)


if __name__ == "__main__":


    # Create a Tk.App(), which handles the windowing system event loop
    root = Tk_get_root()
    # Setting consistent style
    style = ttk.Style()
    style.theme_use('xpnative')
    # Building main player window
    root.iconbitmap(default='./Resources/Tie.ico')
    root.protocol("WM_DELETE_WINDOW", _quit)
    player = Player(root, title="TIE Suite 1.2")
    # show the player window centered and run the application
    root.mainloop()
